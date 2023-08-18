import requests
from playwright.sync_api import sync_playwright
from dataclasses import dataclass, asdict, field
import pandas as pd
import argparse
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup as bs
import json
import csv
from pynput.keyboard import Key, Controller
import psycopg2
import time


class Rieltors:
    def __init__(self):
        self.domain = ''
        self.link_array = []

        self.phone = []
        self.email = []
        self.address = []
        self.city = []
        self.license_r = []
        self.listed_by = []
        self.rieltor_name_array = []

        self.cities_array = []

        self.email_array  = []
        self.phone_array  = []
        self.name_array  = []

        self.state_array = []
        self.city_array = []
        self.zip_array = []

        self.addreses_what_need_check = []
        self.spend_addreses = []
        self.company_array = []
        self.index_for_save = 0

    def make_request(self):
        pass
    def create_file(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['B1'] = "Phones"
        ws['C1'] = "Emails"
        ws['D1'] = "State"
        ws['E1'] = "City"
        ws['F1'] = "ZIP"
        ws['G1'] = "Company"

        wb.save("rieltors.xlsx")
    def save_data(self):

        xlsx_file_path = "rieltors.xlsx"
        wb = load_workbook(xlsx_file_path)

        # Step 2: Select the worksheet where you want to append data
        sheet_name = "Sheet"  # Change this to the name of your target sheet
        sheet = wb[sheet_name]

        # Calculate the next row to append data to (assuming data starts from row 2)
        next_row = sheet.max_row + 1
        for i in range(len(self.name_array)):
            try:
                sheet.cell(row=next_row, column=1, value=str(self.name_array[i]))
                sheet.cell(row=next_row, column=2, value=str(self.phone_array[i]))
                sheet.cell(row=next_row, column=3, value=str(self.email_array[i]))
                sheet.cell(row=next_row, column=4, value=str(self.state_array[i]))
                sheet.cell(row=next_row, column=5, value=str(self.city_array[i]))
                sheet.cell(row=next_row, column=6, value=str(self.zip_array[i]))
                sheet.cell(row=next_row, column=7, value=str(self.company_array[i]))
            except Exception as e:
                print(e)
            next_row += 1
        wb.save(xlsx_file_path)

        # Optional: Close the workbook
        wb.close()





    def get_summary_field(self,page):
        return page.inner_text('div.SearchList__summary', timeout=500000)

    def check_page(self,page,index,zip_code):
        # check if element exists div.KWPropertyCard__courtesy
        page.wait_for_selector('div.KWPropertyCard__courtesy', timeout=20000)

        properties = page.query_selector_all('div.KWPropertyCard__courtesy')
        properties_adreses = page.query_selector_all('div.KWPropertyCardInfo__address')
        print(f"properties_adreses")
        print(len(properties_adreses))
        for property_adres in properties_adreses:
                if property_adres.inner_text() not in self.addreses_what_need_check and property_adres.inner_text() not in self.spend_addreses:
                    self.addreses_what_need_check.append(property_adres.inner_text())


        # loop through each property
        for property_p in properties:
            property_rieltor_name = property_p.inner_text()
            # print(property_rieltor_name)
            # if Keller, EXP or Century 21 in property name
            # if "Keller" in property_rieltor_name or "EXP" in property_rieltor_name or "Century 21" in property_rieltor_name:
            # print(property_rieltor_name)
            # get parent element
            parent = property_p.query_selector("xpath=..")
            # get property link
            link = parent.get_attribute('href')
            # print(link)
            if link not in self.link_array:
                self.link_array.append(link)

        print(f"self.link_array {len(self.link_array)}")
        # get div class SearchList__summary text
        summary_field = self.get_summary_field(page)
        of_p = summary_field.find("of")
        properties_pos = summary_field.find(" Properties")
        # cut summary_field to get number of properties
        properties_count = summary_field[of_p+3:properties_pos]
        print('properties_count')
        # if index == 0:
        try:
            propersties_count_int = int(properties_count.strip())
        except Exception as e:
            print(e)
            propersties_count_int = 1000
        print(f"propersties_count_int {propersties_count_int}")
        if propersties_count_int > 40:
            # click button with button class KWZoomControl__button
            page.click('button.KWZoomControl__button', timeout=500000)
            # just wait for 5 seconds
            page.wait_for_timeout(20000)
            # index += 1
            self.check_page(page, index,zip_code)
        elif propersties_count_int <= 40:
            for addrs in self.addreses_what_need_check:
                # remove from self.addreses_what_need_check
                self.addreses_what_need_check.remove(addrs)
                # add to self.spend_addreses
                self.spend_addreses.append(addrs)
                try:
                    self.page_per_page(page, str(addrs), zip_code)
                except Exception as e:
                    print(e)




        # if page.inner_text('text="See More"', timeout=500000):
        #     # click Show more button
        #     page.click('text="See More"')
        index += 1
        print(f"self.link_array {len(self.link_array)} and index {index}")
        # if index < 2:
        #     self.check_page(page,index)

    def generate_link(self):
        # link sample https://www.kw.com/search/location/ChIJX9vMLpaC54gRoeGj7jz2wek-0.8984511517167186,Florida%20Ocoee%2034761,Ocoee%2C%20FL%2034761%2C%20USA?fallBackCityAndState=Ocoee%2C%20FL&fallBackPosition=28.5830551%2C%20-81.52672820000001&fallBackStreet=&isFallback=true&viewport=28.61479654954222%2C-81.48971430441892%2C28.52615080507333%2C-81.58086649558103&zoom=13
        state = "Florida"
        city = "Aventura"
        zip = "33160"
        state_code = "FL"
        link = f"https://www.kw.com/search/location/ChIJX9vMLpaC54gRoeGj7jz2wek-0.8984511517167186,{state}%20{city}%20{zip},{city}%2C%20FL%20{zip}%2C%20USA?fallBackCityAndState={city}%2C%20{state_code}&fallBackPosition=28.5830551%2C%20-81.52672820000001&fallBackStreet=&isFallback=true&viewport=28.61479654954222%2C-81.48971430441892%2C28.52615080507333%2C-81.58086649558103&zoom=13"
        print(link)

    def page_per_page(self,page,go_to_link_str,zip_code):
        # generate link
        page.goto("https://www.kw.com/", timeout=100000)

        # set data into input with class KWSearchInput__input
        page.fill('input.KWSearchInput__input', go_to_link_str, timeout=100000)

        # click search enter
        page.press('input.KWSearchInput__input', 'Enter')

        # check if text exist See More
        # wait for 10 seconds
        page.wait_for_timeout(10000)
        # find all div elements with class name 'KWPropertyCard__light'
        # this is the container for each property
        try:
            click_index = 0
            self.search_new_links(page,click_index)
            # self.check_page(page, 0, zip_code)
        except Exception as e:
            print(e)

    def search_new_links(self,page,click_index=None):
        page.wait_for_selector('div.KWPropertyCard__courtesy', timeout=20000)

        properties = page.query_selector_all('div.KWPropertyCard__courtesy')
        # loop through each property
        for property_p in properties:
            property_rieltor_name = property_p.inner_text()
            # print(property_rieltor_name)
            # if Keller, EXP or Century 21 in property name
            # if "KELLER" in property_rieltor_name or "CENTURY" in property_rieltor_name or "Keller" in property_rieltor_name or "EXP" in property_rieltor_name or "Century 21" in property_rieltor_name:
            # print(property_rieltor_name)
            # get parent element
            parent = property_p.query_selector("xpath=..")
            # get property link
            link = parent.get_attribute('href')

            if link not in self.link_array:
                self.link_array.append(link)

        print(len(self.link_array))

        # mouse move to the span with text See More
        page.hover('text="Save Search"', timeout=10000)
        # scroll down to several pixels
        index = 0
        index2 = 100
        for i in range(0, 10):

            # scroll down to several pixels
            # using JS script
            page.evaluate("window.scrollBy(0, " + str(index) + ")")
            index += 80
            index2 += 100
        # check if bootom of the page
        # check if \"See More\ button exist

        if page.query_selector('button.KWButton.KWButton--secondary.KWButton--medium.KWButton--block'):
            print(f"click_index {click_index}")
            # click Show more button
            if click_index == 5:
                # click button with class KWButton KWButton--secondary KWButton--medium KWButton--block
                page.click('button.KWButton.KWButton--secondary.KWButton--medium.KWButton--block', timeout=20000)
                # page.click('text="See More"', timeout=10000)
                click_index = 0
            # just wait for 10 seconds
            page.wait_for_timeout(10000)
            click_index += 1
            self.search_new_links(page,click_index)

    def save_step(self):
        for link_param in self.link_array:
            link = "https://www.kw.com/" + str(link_param)
            print(link)
            # meake request to link
            lisk_res = requests.get(link,timeout=10000)
            print(lisk_res.status_code)
            # print(lisk_res.text)
            if lisk_res.status_code == 200:
                bs_result = bs(lisk_res.text, 'html.parser')

                # find script by id '__NEXT_DATA__'
                script = bs_result.find('script', id='__NEXT_DATA__').text

                json_res = json.loads(script)
                # print(json_res['props']['pageProps']['propertyData']['listingAgentData']['courtesyOfBrokerage'])
                # quit()

                json_data = json_res['props']['pageProps']['propertyData']['listingAgentData']
                json_state = json_res['props']['pageProps']['propertyData']['locator']['address']['state']
                json_city = json_res['props']['pageProps']['propertyData']['locator']['address']['city']
                json_zip = json_res['props']['pageProps']['propertyData']['locator']['address']['zipcode']
                json_broker_name = json_data['courtesyOfBrokerage']
                # self.state_array.append(json_state)
                # self.city_array.append(json_city)
                # self.zip_array.append(json_zip)
                # self.company_array.append(
                #     json_res['props']['pageProps']['propertyData']['listingAgentData']['courtesyOfBrokerage'])

                print(json_data)
                try:
                    full_name = json_data['fullName']
                    if full_name == None:
                        full_name = json_data['brokerLicense']
                except:
                    full_name = ''
                print(full_name)
                # if full_name == None:
                #     print(json_res)
                phones_arr = []

                type = ''
                phone = ''
                try:
                    for phones in json_data['phones']:
                        phoneNumber = phones['phoneNumber']
                        phoneNumberType = phones['phoneNumberType']
                        phones_arr.append(str(phoneNumberType) + " " + str(phoneNumber))
                    print(phones_arr)
                except:
                    phones_arr.append('')
                email = ''
                try:
                    for cont_E in json_data['contactMethods']:
                        if "@" in cont_E['value']:
                            email = cont_E['value']

                except:
                    email = ''
                print(email)

                # self.email_array.append(email)
                # self.phone_array.append(" ".join(phones_arr))
                # self.name_array.append(full_name)
                address = ''
                if json_city == None:
                    json_city = ''
                # print(email)
                # print(" ".join(phones_arr))
                # print(full_name)
                # print(json_state)
                # print(json_city)
                # print(json_zip)
                try:
                    self.insert_data_on(email, " ".join(phones_arr), full_name, json_state, json_city, json_zip, json_broker_name)
                except:
                    pass
        self.clean_all_variables()

    def clean_all_variables(self):
        # clear all variables
        for name, value in vars(self).items():
            # clear all variables
            if type(value) is list and name != 'cities_array':
                value.clear()
        print("Data saved")

    def goto(self):
        # start selenium browser
        print(f"self.cities_array {len(self.cities_array)}")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()

            # wait until page is loaded
            index_step = 0
            print(f"self.cities_array {len(self.cities_array)}")
            for go_to_link in self.cities_array:
                zip_code = go_to_link[2]
                go_to_link_str = " ".join(go_to_link)
                # print(type(go_to_link_str))
                print(go_to_link_str)
                try:
                    self.page_per_page(page,go_to_link_str,zip_code)
                except:
                    pass

                index_step += 1
                self.index_for_save += 1
                print(f"index_step {index_step}")
                print(f"self.link_array {len(self.link_array)}")

                # save
                if len(self.link_array) > 1000:
                    print("len(self.link_array) % 10")
                    print(len(self.link_array))
                    try:
                        self.save_step()
                    except Exception as e:
                        print(e)


                # if index_step == 4:
                #     break

            # print(len(self.addreses_what_need_check))








    def read_csv_file(self):
        self.cities_array = []
        # read csv file uscities.csv line by line
        forvard = False
        with open('uscities.csv', newline='') as csvfile:
            # get state_name
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row['city_ascii'] == "Davie":
                    forvard = True
                if forvard == True:
                    if row['state_name'] == 'Florida': #and row['city_ascii'] != 'Miami' and row['city_ascii'] != 'Tampa'
                        zip_arr = row['zips'].split(' ')
                        for zip in zip_arr:
                            # print(row['state_name'],row['city_ascii'],zip)
                            self.cities_array.append([str(row['state_name'])+",",str(row['city_ascii'])+",",str(zip)])

        return self.cities_array


    def postgres_connect(self):
        connection = None
        try:
            # In PostgreSQL, default username is 'postgres' and password is 'postgres'.
            # And also there is a default database exist named as 'postgres'.
            # Default host is 'localhost' or '127.0.0.1'
            # And default port is '54322'.
            postgre_user = 'postgres'
            postgre_host = '64.225.108.120'
            postgre_password = '486070920'
            postgre_port = '5432'
            connection = psycopg2.connect(
                f"user='{postgre_user}' host='{postgre_host}' password='{postgre_password}' port='{postgre_port}'")


            return connection

        except:
            print('Database not connected.')


    def select_all_data_from_data_collection(self):
        connection = self.postgres_connect()
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM data_collections_data_collections LIMIT 10")
        rows = cursor.fetchall()
        # print(rows)
        return rows

    def insert_data_on(self,email, phone, name, state, city, zip_code, address):

        connection = self.postgres_connect()
        cursor = connection.cursor()
        cursor.execute("INSERT INTO data_collections_data_collections (email, phone, name, state, city, zip_code, address) VALUES (%s, %s, %s, %s, %s, %s, %s)", (email, phone, name, state, city, zip_code, address))
        connection.commit()
        print("Record inserted successfully into data_collections_data_collections table")

    def remvoe_duplicate(self):
        connection = self.postgres_connect()
        cursor = connection.cursor()
        # select lasr 10 records
        cursor.execute("SELECT * FROM data_collections_data_collections ORDER BY id DESC")
        rows = cursor.fetchall()
        print(rows)
        for row in rows:
            name = row[3]
            state = row[4]
            city = row[5]
            zip_code = row[6]
            # select all records with same name, state, city
            cursor.execute("SELECT * FROM data_collections_data_collections WHERE name = %s AND state = %s AND city = %s AND zip_code = %s", (name, state, city,zip_code))
            rows = cursor.fetchall()
            # print(rows)
            # len rows
            print(len(rows))
            # loop from 1 element, not 0
            if len(rows):
                for i in range(1,len(rows)):
                    print(rows[i][0])
                    # delete all duplicates
                    cursor.execute("DELETE FROM data_collections_data_collections WHERE id = %s", (rows[i][0],))
                    connection.commit()
            # quit()

            #
            # cursor.execute("DELETE FROM data_collections_data_collections a USING data_collections_data_collections b WHERE a.ctid < b.ctid AND a.email = b.email;")
            # connection.commit()
        # print("Duplicate removed")

    def pape_per_page(self,page,i):
        print('pape_per_page')
        print(f"I step = {i}")
        # try:
            # div.FindAgentRoute__row:nth-child(1)
            # find div class AgentCard__text AgentCard__row 1

        # company = page.inner_text("div.FindAgentRoute__row:nth-child(" + i + ")", timeout=8000)
        # # get only text
        # print(f"company = {company}")
        # company = company


        # click in new tab
        try:
            page.click("div.FindAgentRoute__row:nth-child(" + str(i) + ")", timeout=8000)
        except:
            # evalute js
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            # just wait
            time.sleep(10)
            self.pape_per_page(page, i)
        try:
            acgent_name = page.inner_text("div.AgentContent__name",timeout=1000)
        except:
            acgent_name = ''
        try:
            agent_location = page.inner_text("div.AgentContent__location",timeout=1000)
            location = agent_location.split(",")
        except:
            location = ''
        try:
            email = page.inner_text("a.AgentInformation__factBody",timeout=1000)
        except:
            email = ''

        try:
            phones = page.inner_text("div.AgentInformation__phoneNumbers",timeout=1000)
        except:
            phones = ''

        try:
            soup = bs(page.content(), 'html.parser')
            # print(soup)
            language = soup.find(text="Market Center")
            # get parent tag of language
            language = language.parent
            # get parent tag of language
            language = language.parent
            company = language.text
            company = company.replace("Market Center", "")
        # print(compant)
        # quit()
        # #
        except:
            company = ''


        print(acgent_name)
        print(location)
        print(email)
        print(phones)
        print(company)
        print("====================================="
              )
        self.insert_data_on(email, phones, acgent_name, location[1], location[0], '', company)
        # click back
        kb = Controller()
        # press alt + left together
        kb.press(Key.alt)  # Presses "up" key
        kb.press(Key.left)  # Presses "left" key
        kb.release(Key.alt)  # Releases "up" key
        kb.release(Key.left)  # etc..
        print('back')


    def zillow_requst(self):
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            go_index = False
            for go_to_link in self.cities_array:
                print(go_to_link)
                city = go_to_link.upper()
                if city == 'Vero Beach South':
                    go_index = True
                if go_index == True:
                    try:
                        # city = go_to_link[1].lower()
                        # print(city)
                        # state = go_to_link[0].lower()
                        # print(state)
                        url = f"https://www.kw.com/agent/search/FL/"+city+"/"

                        page.goto(url)
                        page.wait_for_load_state()
                        total_count = page.inner_text("div.FindAgentRoute__totalCount")

                        only_digits = [digit for digit in total_count if digit.isdigit()]
                        only_digits = "".join(only_digits)
                        print(only_digits)
                        # find div with attr
                        # CLICK on div.FindAgentRoute__row with index 1
                        for i in range(1, 50):
                            print(f" I = {i}")
                            try:
                                self.pape_per_page(page,i)
                            except:
                                # page.goto(url)
                                # time.sleep(10)
                                pass
                    except Exception as e:
                        print(e)
                # quit()

                # page.wait_for_timeout(100000)

        # response = requests.get("https://www.realtor.com/realestateagents/56bbecba7e54f7010021baa7",timeout=10)
        #
        # print(response.text)
    def read_csv_file2(self):
        self.cities_array = []
        # read csv file uscities.csv line by line
        with open('uscities.csv', newline='') as csvfile:
            # get state_name
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row['state_name'] == 'Florida':
                    # print(row['city'], row['state_name'])
                    self.cities_array.append(row['city'])
            # quit()

    def save_data_into_excel(self):
        connection = self.postgres_connect()
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM data_collections_data_collections")
        rows = cursor.fetchall()

        index = 1
        index_last = 0
        for row in rows:
            print(row[1])
            self.email_array.append(row[1])
            self.phone_array.append(row[2])
            self.name_array.append(row[3])
            self.state_array.append(row[4])
            self.city_array.append(row[5])
            self.zip_array.append(row[6])
            self.company_array.append(row[7])
            index_last += 1
            if len(self.email_array) == 10000:
                index = self.save_into_excel(index,index_last)
                # clen all arrays
                for name, value in vars(self).items():
                    # clear all variables
                    if type(value) is list:
                        value.clear()
        index = self.save_into_excel(index, index_last)


    def save_into_excel(self,index,index_last):
        print(f"index,index_last {index,index_last}" )
        # create file if not exist
        try:
            # load file
            wb = load_workbook("rieltors"+str(index)+"_"+str(index_last)+".xlsx")
            # get active sheet
            ws = wb.active
            # get last row
            last_row = ws.max_row
            # print(last_row)

            # loop through each element
            for i in range(len(self.name_array)):
                # save data into excel
                ws.cell(row=last_row+i+1, column=1, value=str(self.name_array[i]))
                ws.cell(row=last_row+i+1, column=2, value=str(self.phone_array[i]))
                ws.cell(row=last_row+i+1, column=3, value=str(self.email_array[i]))
                ws.cell(row=last_row+i+1, column=4, value=str(self.state_array[i]))
                ws.cell(row=last_row+i+1, column=5, value=str(self.city_array[i]))
                ws.cell(row=last_row+i+1, column=6, value=str(self.zip_array[i]))
                ws.cell(row=last_row+i+1, column=7, value=str(self.company_array[i]))

            # save file
            wb.save("rieltors"+str(index)+"_"+str(index_last)+".xlsx")
            wb.close()

        except:
            # create file
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Name"
            ws['B1'] = "Phones"
            ws['C1'] = "Emails"
            ws['D1'] = "State"
            ws['E1'] = "City"
            ws['F1'] = "ZIP"
            ws['G1'] = "Company"

            # save file with name
            wb.save("rieltors"+str(index)+"_"+str(index_last)+".xlsx")

            self.save_into_excel(index,index_last)

        return index_last+1




if __name__ == "__main__":
    rieltors = Rieltors()
    rieltors.read_csv_file()
    rieltors.create_file()
    rieltors.domain = "https://www.kw.com/search/location/ChIJY10Hv_i02YgRjdzvoWOVM6w-0.7420868142967443,Florida%2C%20Miami%2C%2033109,Miami%20Beach%2C%20FL%2033109%2C%20USA?fallBackCityAndState=Miami%20Beach%2C%20FL&fallBackPosition=25.7560139%2C%20-80.1344842&fallBackStreet=&isFallback=true&viewport=25.872362435965854%2C-80.1049454695791%2C25.826943802010476%2C-80.15103654929102&zoom=14"
    rieltors.goto()
    #
    # rieltors.remvoe_duplicate()

    # rieltors.read_csv_file2()
    # # rieltors.read_csv_file()
    # rieltors.zillow_requst()

    # rieltors.select_all_data_from_data_collection()
    #
    # rieltors.save_data_into_excel()







