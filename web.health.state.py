# https://hcir.web.health.state.mn.us/searchInterpreter.jsp
import pyautogui
import requests
import random
from playwright.sync_api import sync_playwright
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
from pynput.keyboard import Key, Controller
from bs4 import BeautifulSoup as bs
import os
import re

class Start:
    def __init__(self):
        self.url = "https://hcir.web.health.state.mn.us/search.do?searchAction=searchResult"

        self.js_code = []
        self.surname = []
        self.name = []
        self.rosted_id = []
        self.rosted_exp = []

        self.language = []
        self.phone = []
        self.email = []
        self.a_location = []
        self.available = []
        self.s_r = []
        self.agen_name = []

    def get_data(self):
        pass

    def goto(self):
        response = requests.get(self.url)
        # print(response.status_code)
        # print(response.text)

        # convert to soup object
        soup = bs(response.text, 'html.parser')

        # get all td with class="tableBoldCell"
        tableBoldCells = soup.find_all('td', class_='tableBoldCell')
        index = 0
        for tableBoldCell in tableBoldCells:
            js_code = ""
            try:
                js_code = tableBoldCell.a['href']
                name = tableBoldCell.text
                # print(js_code)
                # print(f"index: {index}")
                # print(f"tableBoldCell.text: {js_code} and name: {name}")
                # print("---------------")
                if index == 0 and len(js_code) > 0:
                    self.js_code.append(js_code)
                    self.surname.append(name)

            except:
                # print(f"index: {index}")
                # print(f"tableBoldCell.text: {tableBoldCell.text}")
                # print("---------------")
                if index == 0 and len(js_code) == 0:
                    self.js_code.append('')
                    self.surname.append(tableBoldCell.text)
                if index == 1:
                    self.name.append(tableBoldCell.text)
                if index == 2:
                    self.rosted_id .append(tableBoldCell.text)
                if index == 3:
                    self.rosted_exp.append(tableBoldCell.text)

            index += 1
            if index == 4:
                index = 0

        print(len(self.js_code))
        print(len(self.surname))
        print(len(self.name))
        print(len(self.rosted_id))
        print(len(self.rosted_exp))


        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            try:
                page.goto(self.url, timeout=60000)
            except:
                self.save_data_to_xlxs()

            step_i = 0
            finded_id = 0
            finded_id_status = False
            for js in self.js_code:
                if self.rosted_id[step_i] == '87528':
                    finded_id = step_i
                    print(f"step_i primary {step_i}")
                    finded_id_status = True
                if finded_id_status == False:
                    self.language.append('')
                    self.phone.append('')
                    self.email.append('')
                    self.a_location.append('')
                    self.available.append('')
                    self.s_r.append('')
                    self.agen_name.append('')
                if finded_id_status == True:
                # if finded_id > 0:
                    try:
                        self.link_by_link(page,js,step_i)
                    except:
                        pass
                step_i += 1
                print(f"step_i: {step_i}")
                if step_i == 1000:
                    break

    def link_by_link(self,page,js,step_i):
        if len(js) == 0:
            self.language.append('')
            self.phone.append('')
            self.email.append('')
            self.a_location.append('')
            self.available.append('')
            self.s_r.append('')
            self.agen_name.append('')

        else:
            try:
                try:
                    # jest wait for iframe to load
                    page.wait_for_timeout(1000)

                    # run javascript javascript:openInterp('12840')
                    print(js)
                    page.evaluate(str(js))
                    page.wait_for_timeout(100)

                    # get full html from page
                    html = page.inner_html('html')

                    self.get_all_data_from_html(html)
                    page.wait_for_timeout(9000)
                    # click Alt + Left Arrow
                    kb = Controller()
                    # press alt + left together
                    kb.press(Key.alt)  # Presses "up" key
                    kb.press(Key.left)  # Presses "left" key
                    kb.release(Key.alt)  # Releases "up" key
                    kb.release(Key.left)  # etc..

                    page.wait_for_timeout(10000)
                except:
                    pass
            except:
                print("Skipping. Connnection error")
                print(f"Broked ID {step_i}")
                page.wait_for_timeout(100000)
                self.link_by_link(page,js,step_i)

    def get_all_data_from_html(self,html=None):
        soup = bs(html, 'html.parser')

        # find data by text "Language(s) and Any Dialect(s):"
        params_arr = {"language":"Language(s) and Any Dialect(s):","phone":"Phone #:",
        "email":"E-Mail Address:",
        "a_location":"Available Locations:",
        "available":"Availability:",
        "s_r":"Specific areas of interest:",
        "agen_name":"Agency Name and Phone #:"}


        for key,param in params_arr.items():
            try:
                language = soup.find(text=str(param))
                # get parent tag of language
                language = language.parent
                # get parent tag of language
                language = language.parent
                # get all data in text
                language = language.text
                language = language.replace(str(param), "")
                language = language.strip()
                if key == "language":
                    self.language.append(language)
                elif key == "phone":
                    self.phone.append(language)
                elif key == "email":
                    self.email.append(language)
                elif key == "a_location":
                    self.a_location.append(language)
                elif key == "available":
                    self.available.append(language)
                elif key == "s_r":
                    self.s_r.append(language)
                elif key == "agen_name":
                    self.agen_name.append(language)
                print(language)
            except:
                print("Not Found")

    def save_data_to_xlxs(self):
        # check if hcir_web_health_state.xlsx exist
        if os.path.isfile("hcir_web_health_state.xlsx") == False:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Last Name"
            ws['B1'] = "First Name"
            ws['C1'] = "Roster ID"
            ws['D1'] = "Roster Expiration Date"
            ws['E1'] = "Language(s) and Any Dialect(s):	"
            ws['F1'] = "Phone #:"
            ws['G1'] = "E-Mail Address:	"
            ws['H1'] = "Available Locations:	"
            ws['I1'] = "Availability:	"
            ws['J1'] = "Specific areas of interest:	"
            ws['K1'] = "Agency Name and Phone #:	"
            ws['L1'] = "Expiration Date:	"
            ws['M1'] = "Roster"

            wb.save("hcir_web_health_state.xlsx")

        xlsx_file_path = "hcir_web_health_state.xlsx"
        wb = load_workbook(xlsx_file_path)

        # Step 2: Select the worksheet where you want to append data
        sheet_name = "Sheet"  # Change this to the name of your target sheet
        sheet = wb[sheet_name]

        # Calculate the next row to append data to (assuming data starts from row 2)



        next_row = sheet.max_row + 1
        print(f"next_row {next_row}")
        for i in range(len(self.surname)):
            try:
                sheet.cell(row=next_row, column=1, value=str(self.surname[i]))
                sheet.cell(row=next_row, column=2, value=str(self.name[i]))
                sheet.cell(row=next_row, column=3, value=str(self.rosted_id[i]))
                sheet.cell(row=next_row, column=4, value=str(self.rosted_exp[i]))
                sheet.cell(row=next_row, column=5, value=str(self.language[i]))
                sheet.cell(row=next_row, column=6, value=str(self.phone[i]))
                sheet.cell(row=next_row, column=7, value=str(self.email[i]))
                sheet.cell(row=next_row, column=8, value=str(self.a_location[i]))
                sheet.cell(row=next_row, column=9, value=str(self.available[i]))
                sheet.cell(row=next_row, column=10, value=str(self.s_r[i]))
                sheet.cell(row=next_row, column=11, value=str(self.agen_name[i]))
                sheet.cell(row=next_row, column=12, value=str(self.rosted_exp[i]))
                sheet.cell(row=next_row, column=13, value=str(self.rosted_id[i]))
                next_row += 1
            except Exception as e:
                print(e)

        wb.save(xlsx_file_path)

        # Optional: Close the workbook
        wb.close()


if __name__ == "__main__":
    start = Start()
    start.goto()
    # # start.get_all_data_from_html()
    start.save_data_to_xlxs()

   #  # openpyxl review cell in hcir_web_health_state.xlsx
   #
   #  wb = load_workbook("hcir_web_health_state.xlsx")
   #
   # # get all F column
   #  sheet = wb.active
   #  for i in range(2,sheet.max_row+1):
   #      # try:
   #      phone_res = sheet.cell(row=i, column=6).value
   #      email_data_res = sheet.cell(row=i, column=7).value
   #      email_data_res_primary = ""
   #
   #
   #      try:
   #          # remove all email and E-Mail Address:
   #          phone_res = phone_res.replace("E-Mail Address:", "")
   #          # find email in string
   #          email_res = re.findall(r'[\w\.-]+@[\w\.-]+', phone_res)
   #          phone_res = phone_res.replace(email_res[0],"")
   #      except:
   #          pass
   #
   #      try:
   #          try:
   #              email_data_res = email_data_res.replace("Phone #:", "")
   #              print(email_data_res)
   #          except:
   #              pass
   #          # find email in string
   #          email_data_res_primary = re.findall(r'[\w\.-]+@[\w\.-]+', email_data_res)
   #          print(f"email_data_res_primary {email_data_res_primary}")
   #      except:
   #          pass
   #
   #      try:
   #          sheet.cell(row=i, column=6, value=phone_res.strip())
   #      except:
   #          sheet.cell(row=i, column=6, value=phone_res)
   #
   #      try:
   #          sheet.cell(row=i, column=7, value=email_data_res_primary[0])
   #      except:
   #          sheet.cell(row=i, column=7, value="")
   #      # save file
   #      wb.save("hcir_web_health_state.xlsx")
   #      # except:
   #      #     pass