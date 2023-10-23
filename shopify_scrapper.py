import time
from bs4 import BeautifulSoup as bs
import requests
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
import json
import sys
import pandas as pd
from playwright.sync_api import sync_playwright
import random
import os
from difflib import SequenceMatcher
import xmltodict

class ShopifyScrapper:

    def __init__(self):
        self.webarchive = False
        self.webarchive_url = ''
        self.webarchive_url_domain = ''
        self.dublicate = []
        self.increment_index = 0
        self.url = ''
        self.id_by_id_arr = list()
        self.id_arr = []
        self.product_name_arr = []
        self.price_arr = []
        self.full_link_arr = []
        self.data_value_list_arr = []
        self.handle_arr = []
        self.related_collections_handle_arr = []
        self.full_description_arr = []
        self.variants_arr = []
        self.full_description_html_arr = []
        self.title_arr = []
        self.title_html_arr = []
        self.ceo_title_arr = []
        self.ceo_description_arr = []
        self.images_arr = []
        self.imge_primary_arr = []
        self.domain = 'https://univers-chinois.com'
        self.https = 'https:'
        self.variants_arr_primary = []
        self.bullet_points_arr = []
        self.h2_html_arr = []
        self.product_id_arr = []
        self.tags_arr = []
        self.vendor_arr = []
        self.type_arr = []
        self.images_arr_variant = []
        self.secure_url_arr = []
        self.published_at = list()
        self.created_at = list()
        self.available_arr = []
        self.compare_at_price_varies_arr = []
        self.price_varies_arr = []
        self.product_counter = []
        self.compare_at_price_arr = []
        self.compare_at_price_max_arr = []
        self.price_min_arr = []
        self.price_max_arr = []
        self.product_hendler = []

        # full decsription arr
        self.total_description_html_arr = []
        self.clean_description_html_arr = []
        self.collection_value = []
        self.hanle = ''


        self.blog_links = []
        self.blog_hendle = []
        self.blog_next_pages = []
        self.blog_ceo_title = []
        self.blog_categories = []
        self.blog_ceo_desc = []
        self.blog_title_text = []
        self.blog_title_html = []
        self.blog_desc_text = []
        self.blog_desc_html = []
        self.blog_feature_image = []
        self.blog_name = ''
        self.extra_blog_pages = []

        self.super_webarchive_collections_links = []
        self.super_webarchive_blog_links  = []
        self.super_webarchive_products_links  = []

        self.primary_collections_site = ''
        self.related_collections_site = ''
        self.variant_price_arr = []

        self.blog_tags_class = 'content-block content-block--small'
        self.blog_div = 'div'

        self.sitemap_link = ''

        self.menu_tag = ''
        self.menu_id = ''
        self.menu_selector_value = ''



    def remove_all_none_tags(self,soup):
        # print('remove_all_none_tags')
        # remove all ul is not None and len(ul.text) > 40
        uls = soup.find_all(['p','h4','ul','ol','table','br'])
        for ul in uls:
            # print(ul)
            if ul is None or len(ul.text) < 15:
                # if ul children the same like parent, then remove
                # print('removed')
                # print(ul)
                ul.decompose()

        return soup

    def cut_full_description(self,soup,full_description):

        fill_description_primary = ''
        bullet_points_arr = []
        related_col_arr = []
        h2_html_origin = ''

        # remove all  attributes data-mce-fragment
        style_tags = soup.find_all(style=True)
        for style_tag in style_tags:
            # remove style attr
            del style_tag['style']
            del style_tag['data-mce-style']
            del style_tag['data-mce-fragment']
            del style_tag['class']
            del style_tag['data-mce-fragment']
            del style_tag['data-mce-selected']
            del style_tag['width']
            del style_tag['border']
            del style_tag['data-sheets-value']
        soup = soup


        full_description_html = str(full_description)
        full_description_html_res = []

        ul_data = bs(full_description_html, 'html.parser')
        # remove all h2
        h2_data = ul_data.find_all('h2')

        for h2 in h2_data:
            h2.decompose()

        ul_data = ul_data
        # print('cut_full_description')
        # print(ul_data)
        # find all ul and p
        ul_data = self.remove_all_none_tags(ul_data)
        print('remove_all_none_tags')
        # print(ul_data)
        ul_data = ul_data.find_all(['p', 'h4','div', 'ul', 'ol', 'table','img'])
        ul_index = 0
        tag_type = ''
        tag_value = ''
        description_status = False
        count_of_tags = len(ul_data)
        tga_index = 0
        print(f"count_of_tags {count_of_tags}")
        not_iquel_status = False
        for ul in ul_data:

            if ul is not None and len(ul.text) > 5:
                # print(ul.name)
                # print(ul)

                tag_name = ul.name
                if tag_type == '':
                    tag_type = tag_name

                if tag_type == tag_name:
                    # print('Iquil')
                    # print(tag_value)
                    # quit()
                    # tag_value += str(ul)
                    if tag_value.find(str(ul)) == -1:
                        tag_value +=  str(ul) #str(full_description_html_res[-1]) +
                    # full_description_html_res.pop(-1)
                    # full_description_html_res.append(tag_value)
                    # print(tag_value)
                    # full_description_html_res[-1] = str(full_description_html_res[-1]) + str(ul)
                    if ul_index == len(ul_data) - 1:
                        # print(f"Last index")
                        # print(tag_type)
                        # print(f"description_status {description_status}")
                        if tag_type == 'h4' and description_status == False or tag_type == 'p' and description_status == False:
                            full_description_html_res.insert(0, str(tag_value))
                            description_status = True
                        else:
                            full_description_html_res.append(tag_value)


                elif tag_type != tag_name:
                    # print('Not Iquil')
                    # print(ul)
                    # print(tag_type)
                    if tag_type == 'h4' and description_status == False or tag_type == 'p' and description_status == False:
                        full_description_html_res.insert(0, str(tag_value))
                        description_status = True
                    else:
                        full_description_html_res.append(tag_value)

                    # reinstall values
                    tag_type = tag_name
                    # tag_value = ''
                    tag_value = str(ul)

                    if ul_index == len(ul_data) - 1:
                        # print(f"Last index")
                        # print(tag_type)
                        # print(f"description_status {description_status}")
                        if tag_type == 'h4' and description_status == False or tag_type == 'p' and description_status == False:
                            full_description_html_res.insert(0, str(tag_value))
                            description_status = True
                        else:
                            full_description_html_res.append(tag_value)

            ul_index += 1

        try:
            fill_description_primary = full_description_html_res[0]
        except Exception as e:
            fill_description_primary = full_description_html
            print(fill_description_primary)
            print(e)
            print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
        # quit()



        bullet_points_arr.clear()

        # full_description_html_res remove first element
        full_description_html_step = full_description_html_res[1:]
        for elem in full_description_html_step:
            bullet_points_arr.append(elem)
        # print(soup)
        try:
            related_col_arr = []
            arr1 = []
            arr1.append('related_collections')
            related_col_arr.append(arr1)
            # related_collections = soup.find('div', class_='product_links').find_all('a')
            # for related_collection in related_collections:
            #     # print(related_collection)
            #     related_collection = related_collection['href']
            #     handle, collection_handele = self.get_handle_and_collection_handle(related_collection)
            #     # print(handle)
            #     related_col_arr.append(handle)
        except Exception as e:
            print(e)
            related_col_arr.append('')
            print(e)
            print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))


        if self.domain == 'https://univers-chinois.com':
            related_col_arr = []
            fill_description_primary = ''
            full_description_html = ''
            ss = bs(full_description, 'html.parser')

            full_style_tags = ss.find_all(style=True)
            for style_tag2 in full_style_tags:
                # remove style attr
                del style_tag2['class']
                del style_tag2['data-mce-fragment']
                del style_tag2['data-mce-selected']
                del style_tag2['width']
                del style_tag2['border']
                del style_tag2['data-sheets-value']
            ss = ss

            # find all p tags
            p_tags = ss.find_all('p')
            index = 0
            for p_tag in p_tags:
                # if p tag has parent div
                if index < 2:
                    fill_description_primary += str(p_tag)
                index += 1

            try:

                related_collections = soup.find('div', class_='product-links').find_all('a')
                for related_collection in related_collections:
                    related_collection = related_collection['href']
                    handle, collection_handele = self.get_handle_and_collection_handle(related_collection)
                    related_col_arr.append(handle)
            except:
                related_col_arr = []


        # print(fill_description_primary)
        # print(bullet_points_arr)
        # print(related_col_arr)
        # quit()

        return fill_description_primary,bullet_points_arr,related_col_arr

    def clena_bad_tags(self,soup):
        style_tags = soup.find_all()
        for style_tag in style_tags:
            # remove style attr
            del style_tag['class']
            del style_tag['data-mce-fragment']
            del style_tag['data-mce-selected']
            del style_tag['width']
            del style_tag['border']
            del style_tag['style']
            del style_tag['data-mce-style']
            del style_tag['data-sheets-value']
            # del style_tag['src']
            # del style_tag['alt']
            del style_tag['data-mce-src']


        return soup

    def remove_all_css_style(self,full_description_html_primary):

        # remove all style tags
        soup = bs(full_description_html_primary, 'html.parser')
        # find all tags

        full_description_html_primary = str(soup)
        return full_description_html_primary


    def get_collections_related(self,title,soup_item,real_soup,product_data):
        title_arr = title.split('<br>')
        # print("+================")
        # print(title_arr[0])
        # print("+================")
        primary_collections = ''
        related_collections = ''
        # find 'Catégories' text in soup_item
        menu_tag = str(self.menu_tag)
        cat_text = real_soup.find('div', role='center')
        print(f"cat_text {cat_text}")
        # if product_data['description'].find('collections') != -1:
        #     print("product_data['description'].find('collections')")
        #     # get all links from product_data['description']
        #     description = product_data['description']
        #     # str to soup
        #     description = bs(description, 'html.parser')
        #     # find all a
        #     all_a = description.find_all('a')
        #     for a in all_a:
        #         href = a.get('href')
        #         # print(href)
        #         if href != None:
        #             if href.find('/collections/') != -1:
        #                 # print(href)
        #                 # cut collection from href
        #                 c_p = href.find('/collections/')
        #                 p_p = href.find('/product')
        #                 collection = href[c_p+13:p_p]
        #                 # print(f"Real collection {collection}")
        #                 if len(primary_collections) == 0:
        #                     primary_collections = str(collection)
        #                 if len(related_collections) == 0:
        #                     related_collections += str(collection)
        #                 else:
        #                     related_collections += ","+str(collection)
        if cat_text != None:
            # find all a
            cat_text = cat_text.find_all('a')
            for text in cat_text:
                # get href
                href = text.get('href')
                # split
                href_arr = href.split('/')[-1]
                if len(href_arr) > 2:
                    if len(primary_collections) == 0:
                        primary_collections = str(href_arr)
                    if len(related_collections) == 0:
                        related_collections += str(href_arr)
                    else:
                        related_collections += ","+str(href_arr)
            # print(related_collections)
        # elif real_soup.find('center') != None:
        #     # get center text in soup_item
        #     center_text = real_soup.find('center')
        #     print(center_text)
        #     # divade by <br/>
        #     # get html text
        #     html_txt = str(center_text)
        #     html_txt = html_txt.replace('<center>','').replace('</center>','')
        #     html_txt = html_txt.split('<br/>')
        #     print(html_txt)
        #     # split by <br>
        #     center_text = center_text.text.split('<br/>')
        #     # for ind,el in enumerate(html_txt):
        #     lowwer = html_txt[0].strip().lower()
        #     lowwer = lowwer.replace(' ', '-')
        #     html_txt[0] = lowwer
        #     # lowwer = html_txt[0].lower()
        #
        #     print(center_text)
        #     print(html_txt)
        #     primary_collections = html_txt[0]
        #     related_collections = str(html_txt[0])

        else:
            div = soup_item.find('ul', role='menubar')
            all_a = div.find_all('a')

            diff_percent_arr = {}
            for a in all_a:
                if a != None:
                    href = a.get('href')
                    a_text = a.text
                    if href is not None:
                        if href.find('/collections/') != -1 or href.find('/pages/') != -1:
                            # print(href)
                            # print(a_text)
                            percent = SequenceMatcher(None, title_arr[0], a_text).ratio()
                            diff_percent_arr[href] = percent
                            # quit()
            # print(diff_percent_arr)
            try:
                primary_collection = max(diff_percent_arr, key=diff_percent_arr.get)
                primary_collections = self.clean_collections(primary_collection)

            except Exception as e:
                print(e)
                print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))


            print(f"primary_collections {primary_collections}")
            # find primary_collection in soup
            for aa in all_a:
                if aa is not None:
                    # print(aa)
                    try:
                        if aa.get('href') == primary_collection:
                            # find ul data
                            related_collections = self.find_ul_data(aa, related_collections)
                            break
                    except:
                        print(f"some err")

        return primary_collections,related_collections

    def cuto_to_cdn(self,img)->str:
        cdn_pos = img.find('//cdn.')
        return img[cdn_pos-5:]

    def request_link_by_link(self,link_by_item,proxy_index=None,s=None):
        # link_by_item = "https://traditions-de-chine.com/collections/bols-chinois/products/bol-chinois-noir"
        print(f"link_by_item {link_by_item}")

         # make request
        response_item = self.make_request(link_by_item, proxy_index, s)
        # print('converrt into soup html')



        # converrt into soup html
        # print(response_item)
        if response_item != False and response_item != None:
            soup_item = bs(response_item.text, 'html.parser')
            just_text = response_item.text
            real_soup = bs(response_item.text, 'html.parser')

            # remove bad tags
            print("clena_bad_tags")
            soup_item = self.clena_bad_tags(soup_item)

            # get attributes data-product
            try:
                product_data = ''
                print('sizeChartsRelentless.product')
                print(len(soup_item.find_all('script')))
                if just_text.find('sizeChartsRelentless.produc') != -1:
                    print('sizeChartsRelentless.product')
                    all_script = soup_item.find_all('script')
                    for script in all_script:
                        # print(script)
                        if script is not None:
                            scr = script.text
                            # print(scr.find('sizeChartsRelentless.product'))
                            if scr.find('sizeChartsRelentless.product = ') != -1:
                                # print(f"scr.find True")
                                s_pos = scr.find('sizeChartsRelentless.product')
                                e_pos = scr.find('sizeChartsRelentless.productCollection')
                                # print(s_pos)
                                scr = scr.strip()
                                product_data =  scr[s_pos+30:e_pos-3]
                                # product_data = product_data.text
                                # print(product_data)
                                break

                elif soup_item.find('script', id='ProductJson-product-template') != -1:
                    print('ProductJson-product-template')
                    product_data = soup_item.find('script', id='ProductJson-product-template')
                    product_data = product_data.text

                elif str(soup_item).find('var meta = {"product":') != -1:
                    print('var meta = {"product":')
                    pos1 = str(soup_item).find('var meta = {"product":')
                    strind_text = str(soup_item)[pos1+22:]

                    pos2 = strind_text.find('};')
                    strind_text = strind_text[:pos2]
                    product_data = strind_text.strip()
                    print(product_data)
                    product_data = product_data.split('},"')
                    product_data = product_data[0]
                    product_data = product_data+'}'

                else:
                    print('product_form')
                    product_data = soup_item.find('div', class_='product_form')
                    product_data = product_data['data-product']
                    product_data = product_data.text
            except Exception as e:
                print(e)
                print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                product_data = False

            if product_data != False:
                images_arr = []
                # try:
                product_data = json.loads(product_data)



                produc_id = product_data['id']
                img_extra = soup_item.find('meta', {'property': 'og:image'})['content']
                try:
                    og_price_amount = soup_item.find('meta', {'property': 'og:price:amount'})['content']
                except:
                    og_price_amount = ''
                # img_extra = self.cuto_to_cdn(img_extra)
                print(f"img_extra {img_extra}")
                print(produc_id)
                try:
                    product_title = product_data['title']
                except:
                    product_title = soup_item.find('h1').text
                finally:
                    # meta property og:title
                    product_title = soup_item.find('meta', {'property': 'og:title'})['content']

                try:
                    product_handle = product_data['handle']
                except:
                    product_handle = link_by_item.split('/')[-1]

                try:
                    published_at = product_data['published_at']
                    created_at = product_data['created_at']
                except:
                    published_at = ''
                    created_at = ''
                try:
                    vendor = product_data['vendor']
                    product_type = product_data['type']
                    tags = product_data['tags']
                except:
                    vendor = ''
                    tags = ''
                    product_type = ''

                variants_arr = []
                secure_url = ''
                full_description_html_primary = ''
                bullet_points_arr = []
                primary_collections = ''
                related_collections = ''
                # try:
                primary_collections,related_collections = self.get_collections_related(product_title,soup_item,real_soup,product_data)
                print(primary_collections,related_collections)
                # except Exception as e:
                #     print(e)
                #     print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                # quit()
                try:
                    full_description = product_data['description']
                except:
                    full_description = soup_item.find('meta', {'property': 'og:description'})['content']

                full_description = str(self.clena_bad_tags(bs(full_description, 'html.parser')))
                total_description_html_arr = full_description


                all_desc = bs(full_description, 'html.parser')

                full_description_html_primary = ''
                bullet_points_arr = []
                related_col_arr = []
                r = []
                related_col_arr.append(r)


                # find all a
                all_a = soup_item.find_all('a')
                collection = ''
                for a in all_a:
                    href = a.get('href')
                    # print(f"Collection {href}")
                    if href != None:
                        if href.find('/collections/') != -1 and href.find(str(product_handle)) != -1:
                            print(href)
                            # cut collection from href
                            c_p = href.find('/collections/')
                            p_p = href.find('/product')
                            collection = href[c_p+13:p_p]
                            print(f"Real collection {collection}")
                            break
                # print(collection)
                try:
                    # print('a')
                    full_description_html_primary, bullet_points_arr, related_col_arr = self.cut_full_description(soup_item,full_description)
                except Exception as e:
                    # display a line of error
                    print(e)
                    print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))

                index = 0
                h2_html_origin = ''
                for h2 in all_desc.find_all('h2'):
                    if index == 0:
                        # remove style attr from h2
                        del h2['style']
                        del h2['class']
                        # remove all css style
                        h2_html_origin = str(h2)

                    h2.decompose()
                    index += 1
                    break


                full_description = bs(full_description, 'html.parser').text

                # get title from head
                ceo_title = soup_item.find('title').text
                ceo_title = ceo_title.strip()
                try:
                    ceo_description = soup_item.find('meta', {'name': 'description'})['content']
                except:
                    ceo_description = soup_item.find('meta', {'property': 'og:description'})['content']


                try:
                    images = product_data['images']
                    for img in images:
                        images_arr.append(img)
                except:

                    images_arr.append(img_extra)


                title_html = soup_item.find('h1')

                data_value_list = []

                id_by_id = product_data['id']
                vendor = product_data['vendor']
                type = product_data['type']
                try:
                    tags = product_data['tags']
                except:
                    tags = ''

                for product in product_data['variants']:
                    if product['id'] not in self.dublicate:
                        try:
                            # print(product)
                            # print('//////////////')
                            secure_url = ''
                            self.dublicate.append(product['id'])

                            self.primary_collections_site = primary_collections
                            if primary_collections != related_collections:
                                if len(related_collections) > 2:
                                    self.related_collections_site = primary_collections+","+related_collections
                                else:
                                    self.related_collections_site = primary_collections
                            else:
                                self.related_collections_site = related_collections

                            # print(f"===============================")
                            # print(self.primary_collections_site)
                            # print(self.related_collections_site)
                            # print(f"===============================")
                            self.variant_price_arr.append(self.cut_compare_price(product['price']))

                            self.product_hendler.append(product_handle)
                            self.id_by_id_arr.append(product['id'])
                            self.product_name_arr.append(product_title)
                            self.total_description_html_arr.append(total_description_html_arr)
                            self.clean_description_html_arr.append(self.remove_all_css_style(full_description_html_primary))
                            self.full_description_html_arr.append(self.remove_all_css_style(full_description_html_primary))
                            try:
                                self.price_arr.append(self.cut_compare_price(product_data['price']))
                            except:
                                self.price_arr.append(og_price_amount)
                            try:
                                self.price_min_arr.append(self.cut_compare_price(product_data['price_min']))
                                self.price_max_arr.append(self.cut_compare_price(product_data['price_max']))
                            except:
                                self.price_min_arr.append(og_price_amount)
                                self.price_max_arr.append(og_price_amount)

                            self.full_link_arr.append(link_by_item)
                            self.data_value_list_arr.append(data_value_list)
                            self.variants_arr.append(','.join(variants_arr))
                            self.collection_value.append(collection)
                            # title section
                            self.title_arr.append(product_title)
                            self.title_html_arr.append(title_html)
                            self.ceo_title_arr.append(ceo_title)
                            self.ceo_description_arr.append(ceo_description)
                            self.full_description_html_arr.append(full_description_html_primary)
                            self.bullet_points_arr.append(bullet_points_arr)
                            try:
                                secure_url = 'https:' + str(product['featured_image']['src'])
                            except:
                                secure_url = img_extra

                            variants = []
                            try:
                                if product['option1'] == None:
                                    variants.append('')
                                else:
                                    variants.append(product['option1'])
                            except:
                                pass

                            try:
                                if product['option2'] == None:
                                    variants.append('')
                                else:
                                    variants.append(product['option2'])
                            except:
                                pass

                            try:
                                if product['option3'] == None:
                                    variants.append('')
                                else:
                                    variants.append(product['option3'])
                            except:
                                pass

                            self.related_collections_handle_arr.append(product_handle)

                            hh = []
                            hh.append(product_handle)
                            # print(f"product_handle {product_handle}")
                            self.handle_arr.append(hh)
                            self.full_description_arr.append(full_description)

                            try:
                                self.imge_primary_arr.append(images_arr[0])
                            except:
                                self.imge_primary_arr.append(img_extra)
                            try:
                                self.images_arr_variant.append(images_arr[0])
                            except:
                                self.images_arr_variant.append(img_extra)

                            self.images_arr.append(','.join(images_arr))
                            self.variants_arr_primary.append(variants)
                            self.h2_html_arr.append(str(self.remove_all_css_style(h2_html_origin)))
                            self.product_id_arr.append(id_by_id)
                            self.tags_arr.append(','.join(tags))
                            self.vendor_arr.append(vendor)
                            self.type_arr.append(type)
                            self.secure_url_arr.append(secure_url)
                            try:
                                self.published_at.append(product_data['published_at'])
                                self.created_at.append(product_data['created_at'])
                                self.available_arr.append(product_data['available'])
                                self.compare_at_price_varies_arr.append(product_data['compare_at_price_varies'])
                                self.price_varies_arr.append(product_data['price_varies'])
                                # self.compare_at_price_varies_arr.append(product_data['compare_at_price_varies'])
                                # self.price_varies_arr.append(product_data['price_varies'])

                                if str(product['compare_at_price']) != 'None':
                                    # print(f"product_data['compare_at_price'] {product['compare_at_price']}")
                                    self.compare_at_price_arr.append(
                                        self.cut_compare_price(product['compare_at_price']))
                                else:
                                    self.compare_at_price_arr.append(product['compare_at_price'])

                                if product_data['compare_at_price_max'] == 'None':
                                    self.compare_at_price_max_arr.append(product_data['compare_at_price_max'])
                                else:
                                    self.compare_at_price_max_arr.append(self.cut_compare_price(product_data['compare_at_price_max']))

                            except:

                                self.published_at.append('')
                                self.created_at.append('')
                                self.available_arr.append('TRUE')
                                self.compare_at_price_varies_arr.append(og_price_amount)
                                self.price_varies_arr.append(og_price_amount)
                                # self.compare_at_price_varies_arr.append('FALSE')
                                # self.price_varies_arr.append(og_price_amount)
                                self.compare_at_price_arr.append(og_price_amount)
                                self.compare_at_price_max_arr.append('0')

                        except Exception as e:
                            print(e)
                            print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                # except Exception as e:
                #     print(e)
                #     print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))




    def cut_compare_price(self,compare_at_price):

        if compare_at_price != 0 and len(str(compare_at_price)) > 2:
            a = str(compare_at_price)
            pos2 = a[len(a) - 2:]
            pos1 = a[:len(a) - 2]
            compare_at_price = str(pos1) + ',' + (pos2)

        return compare_at_price

    def save_to_xlsx_product_count(self):

            xlsx_file_path = "shopify.xlsx"
            wb = load_workbook(xlsx_file_path)

            # Step 2: Select the worksheet where you want to append data
            sheet_name = "Sheet"  # Change this to the name of your target sheet
            sheet = wb[sheet_name]

            # Calculate the next row to append data to (assuming data starts from row 2)
            next_row = sheet.max_row + 1
            print(self.increment_index)
            sheet.cell(row=next_row, column=1, value="Product Quantity:"+str(self.increment_index))

            wb.save(xlsx_file_path)
            wb.close()


    def save_to_xlsx(self):
        xlsx_file_path = "shopify.xlsx"
        wb = load_workbook(xlsx_file_path)

        # Step 2: Select the worksheet where you want to append data
        sheet_name = "Sheet"  # Change this to the name of your target sheet
        sheet = wb[sheet_name]

        # Calculate the next row to append data to (assuming data starts from row 2)
        next_row = sheet.max_row + 1

        for i in range(len(self.id_by_id_arr)):
            try:
                option1 = self.variants_arr_primary[i][0]
            except:
                option1 = ''
            try:
                option2 = self.variants_arr_primary[i][1]
            except:
                option2 = ''
            try:
                option3 = self.variants_arr_primary[i][2]
            except:
                option3 = ''

            try:
                bullet_points_variant1 = self.remove_all_css_style(self.bullet_points_arr[i][0])
            except:
                bullet_points_variant1 = ''

            try:
                bullet_points_variant2 = self.remove_all_css_style(self.bullet_points_arr[i][1])
            except:
                bullet_points_variant2 = ''

            try:
                bullet_points_variant3 = self.remove_all_css_style(self.bullet_points_arr[i][2])
            except:
                bullet_points_variant3 = ''
            #
            # try:
            #     # remove all style attribute
            #     total_description_html = ''
            #     total_description_html = str(h2_html_arr[i][0]) + str(full_description_html_arr[i][0]) + str(bullet_points_variant1) + str(bullet_points_variant2) + str(bullet_points_variant3)
            #
            # except:
            #     total_description_html = ''

            # Step 4: Append the data to the selected worksheet
            # print(self.handle_arr[i])
            try:
                sheet.cell(row=next_row, column=1, value=str(self.id_by_id_arr[i]))
                sheet.cell(row=next_row, column=2, value=str(self.product_id_arr[i]))
                sheet.cell(row=next_row, column=3, value=str(self.full_link_arr[i]))
                sheet.cell(row=next_row, column=4, value=str(self.product_hendler[i]))
                sheet.cell(row=next_row, column=5, value=str(self.primary_collections_site))
                sheet.cell(row=next_row, column=6, value=str(self.related_collections_site))# self.related_collections_handle_arr[i])
                sheet.cell(row=next_row, column=7, value=str(self.title_arr[i]))
                sheet.cell(row=next_row, column=8, value=str(self.title_html_arr[i]))
                sheet.cell(row=next_row, column=9, value=str(self.ceo_title_arr[i]))
                sheet.cell(row=next_row, column=10, value=str(self.ceo_description_arr[i]))
                sheet.cell(row=next_row, column=11, value=str(self.product_name_arr[i]))
                sheet.cell(row=next_row, column=12, value=str(self.full_description_arr[i]))
                sheet.cell(row=next_row, column=13, value=str(self.total_description_html_arr[i])).fill = PatternFill(start_color='ADFF2F', end_color='ADFF2F', fill_type='solid')
                sheet.cell(row=next_row, column=14, value=str(self.h2_html_arr[i])).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                sheet.cell(row=next_row, column=15, value=str(self.clean_description_html_arr[i])).fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6', fill_type='solid')
                sheet.cell(row=next_row, column=16, value=str(bullet_points_variant1)).fill = PatternFill(start_color='CD853F', end_color='CD853F', fill_type='solid')
                sheet.cell(row=next_row, column=17, value=str(bullet_points_variant2)).fill = PatternFill(start_color='FFEBCD', end_color='FFEBCD', fill_type='solid')
                sheet.cell(row=next_row, column=18, value=str(bullet_points_variant3)).fill = PatternFill(start_color='FF7F50', end_color='FF7F50', fill_type='solid')
                sheet.cell(row=next_row, column=19, value=str(self.published_at[i]))
                sheet.cell(row=next_row, column=20, value=str(self.created_at[i]))
                sheet.cell(row=next_row, column=21, value=str(self.vendor_arr[i]))
                sheet.cell(row=next_row, column=22, value=str(self.type_arr[i]))
                sheet.cell(row=next_row, column=23, value=str(self.tags_arr[i]))
                sheet.cell(row=next_row, column=24, value=str(self.price_arr[i]))
                sheet.cell(row=next_row, column=25, value=str(self.price_min_arr[i]))
                sheet.cell(row=next_row, column=26, value=str(self.price_max_arr[i]))
                sheet.cell(row=next_row, column=27, value=str(self.available_arr[i]))
                sheet.cell(row=next_row, column=28, value=str(self.price_varies_arr[i]))
                sheet.cell(row=next_row, column=29, value=str(self.compare_at_price_arr[i]))
                sheet.cell(row=next_row, column=30, value=str(self.compare_at_price_max_arr[i]))
                sheet.cell(row=next_row, column=31, value=str(self.compare_at_price_varies_arr[i]))
                sheet.cell(row=next_row, column=32, value=str(''))
                sheet.cell(row=next_row, column=33, value=str(''))
                sheet.cell(row=next_row, column=34, value=str(self.remove_webarchive_from_img(self.images_arr[i])))
                sheet.cell(row=next_row, column=35, value=str(self.remove_webarchive_from_img(self.imge_primary_arr[i])))
                sheet.cell(row=next_row, column=36, value=str(self.variants_arr[i]))
                sheet.cell(row=next_row, column=37, value=str(option1))
                sheet.cell(row=next_row, column=38, value=str(option2))
                sheet.cell(row=next_row, column=39, value=str(option3))
                sheet.cell(row=next_row, column=40, value=str(self.remove_webarchive_from_img(self.secure_url_arr[i])))
                sheet.cell(row=next_row, column=41, value=str(''))
                sheet.cell(row=next_row, column=42, value=str(self.variant_price_arr[i]))
                next_row += 1
            except:
                pass


        # Step 5: Save the changes back to the XLSX file
        wb.save(xlsx_file_path)

        # Optional: Close the workbook
        wb.close()
        self.increment_index += len(dict.fromkeys(self.product_id_arr))

        # clear
        # get all variables from __init__
        for name, value in vars(self).items():
            # clear all variables
            if type(value) is list and name != 'super_webarchive_collections_links' and name != 'super_webarchive_blog_links' and name != 'super_webarchive_products_links':
                value.clear()

    def remove_webarchive_from_img(self, img):
        # print('remove_webarchive_from_img')
        # print(img)
        img_arr = []
        if img != '':
            if self.webarchive == True:
                img_arr = img.split(',')
                for i,im in enumerate(img_arr):
                    # print(im)
                    end_pos = im.find('//cdn')
                    # print(end_pos)
                    img_arr[i] = im[end_pos+2:]
                    # print(img_arr[i])

                return ",".join(img_arr)
            elif self.webarchive == False:
                img_arr = img.split(',')
                for i, im in enumerate(img_arr):
                    if im.find('https:') == -1:
                        img_arr[i] = 'https:' + im
                # print(img_arr)
                return ",".join(img_arr)
        else:
            # print(img_arr)
            return img

    def save_to_csv(self,id_by_id_arr,product_name_arr, price_arr,full_link_arr,
                    data_value_list_arr,variants_arr,related_collections_handle_arr,handle_arr,
                    full_description_arr,full_description_html_arr,title_arr,title_html_arr,ceo_title_arr,
                    ceo_description_arr,images_arr,imge_primary_arr):
        import csv
        with open('shopify.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["id","full_link","handle","collection_handele","related_collections_handle","title","title_html","ceo_title","ceo_description","product_name","full_description","full_description_html","h2_html","bullet_points_html","published_at","created_at","vendor","type","tags","price","price_min","price_max","available","price_varies","compare_at_price","compare_at_price_max","compare_at_price_varies","requires_selling_plan","selling_plan_groups","images","featured_image","variants","variant featured_image","variant compare_at_price","variant price"])
            for i in range(len(product_name_arr)):
                writer.writerow([str(id_by_id_arr[i]),str(full_link_arr[i]),str(handle_arr[i]),str(related_collections_handle_arr[i]),str(related_collections_handle_arr[i]),str(title_arr[i]),str(title_html_arr[i]),str(ceo_title_arr[i]),str(ceo_description_arr[i]),str(product_name_arr[i]), str(full_description_arr[i]),str(full_description_html_arr[i]),'','','','','','','',str(price_arr[i]),str(price_arr[i]),str(price_arr[i]),"TRUE","0","","","0","","",str(images_arr[i]),str(imge_primary_arr[i]),str(variants_arr[i][0]),str(images_arr[i]),"",str(price_arr[i])])

    def get_handle_and_collection_handle(self,link):
        # remove all after ? in link
        print(link)
        hendle = link.split('/')[-1]
        collection_handle = link.split('/')[-3]
        return hendle,collection_handle

    def request_link_by_link_to_get_ids(self,link,proxy,s):
        # request link
        response = self.make_request(link,proxy,s)
        if response != None:
            soup = bs(response.text, 'html.parser')

            # get div with class swatch_options and find all data-id and append to id_arr
            id_arr = []
            variants_arr = []
            #
            scripts = soup.find_all('script')
            for script in scripts:
                # find var meta
                if script.text.find('var meta') > -1:
                    # cut string
                    # find { from start and }}; from end
                    script_text = script.text[script.text.find('var meta = {') + 10:script.text.rfind('}};') + 2]
                    print("script_text")
                    # convert string to json
                    script_json = json.loads(script_text)
                    id_by_ids = script_json['product']['variants']
                    for id_by_id in id_by_ids:
                        id_arr.append(id_by_id['id'])
            for div in soup.find_all('div', class_='swatch_options'):
                for variant in div.find_all('div', class_='option_title'):
                    # if div has text, append to handle_arr
                    if len(div.text) > 0:
                        variants_arr.append(variant.text)
            return id_arr,variants_arr

    def make_request(self,url,proxy,s,reconnect=None):
        response = ''
        timeout = 1000
        try:
            if reconnect == True:
                timeout = 100
            response = requests.get(url, timeout=timeout)
        except:
            print('Except')
            print(url)
            print(response)
            self.make_request(url,proxy,s,reconnect=True)

            # response = requests.get(url, timeout=20)
            # print(f"make_request {response.status_code}")
        try:
            if response.status_code == 200:
                return response
            if response.status_code == 404:
                return False
            else:
                time.sleep(60)
                return self.make_request(url,proxy,s)
        except:
            pass


    def cut_collection_link(self,link):
        if "/collections/" in link:
            col_p = link.find('/collections/')
            prod_pos = link.find('/products/')
            self.hanle = link[col_p+13:]
            print("====================================")
            print(self.hanle)
            print("====================================")

    def scrap_shopify(self,all_categpries):
        #super_webarchive_products_links
        if self.webarchive == True:
            domain = self.webarchive_url_domain
        else:
            domain = self.domain

        if len(self.super_webarchive_products_links) > 0:
            index = 0
            for fill_link in self.super_webarchive_products_links:
                # try:
                print(f"Link origin {fill_link}")
                if self.webarchive == True:
                    fill_link = self.webarchive_url_domain + fill_link
                elif self.webarchive == False:
                    fill_link = fill_link
                self.request_link_by_link(fill_link)

                print("++++++++++++")
                print(f'INDEX {index}')
                print(f"ID COUNT {len(self.product_counter)}")
                print(f"handle = {self.hanle}")

                # remove duplicates from list
                print("++++++++++++")
                # if index % 2 == 0:
                try:
                    self.save_to_xlsx()
                except Exception as e:
                    print(e)
                    print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                index += 1
                print(f"Prim INDEX = {index}")
                # if index == 20:
                #     break

        elif len(all_categpries) > 0:
            index = 0
            for category in all_categpries:
                if "/collections/" in category:
                    self.cut_collection_link(category)

                    url = domain + category
                    print(f"url for collection {url}")
                    ip_addresses = [
                        "173.176.14.246:3128",
                        "129.153.157.63:3128",
                        "141.101.115.2:80",
                        "172.67.34.58:80",
                        "172.67.177.251:80",
                        "203.22.223.150:80"
                    ]
                    try:
                        proxy_index = random.choice(ip_addresses)
                        proxy = {"http": proxy_index}
                        s = requests.Session()
                        response = self.make_request(url,proxy,s)
                    except:
                        response = requests.get(url, timeout=200, headers={'User-Agent': 'Mozilla/5.0'})
                        print("Skipping. Connnection error")

                    # request link with proxy
                    if response != False and response != None:

                        soup = bs(response.text, 'html.parser')

                        # find all 'a' with class product-info__caption and get href
                        for link in soup.find_all('a'): # , class_='product-info__caption'
                            link = link.get('href')
                            fill_link = ''

                            if link != None:

                                if self.webarchive == True:
                                    # print(link)
                                    if link.find("/products/") != -1 and "/collections/" not in link:
                                        if "web.archive" in link:
                                            # fill_link = self.cut_collection_link(link)
                                            fill_link = link
                                            print(f"fl {fill_link}")
                                        else:
                                            fill_link = domain + link
                                            print(f"fl {fill_link}")

                                if self.webarchive == False:

                                    if link.find("/products/") != -1 and "/collections/" in link:
                                        fill_link = domain + link
                                        print(f"fl {fill_link}")
                                    # fill_link = ''

                                if len(fill_link) > 0:
                                    # try:
                                        print(f"Link origin {fill_link}")
                                        self.request_link_by_link(fill_link,proxy,s)

                                        print("++++++++++++")
                                        print(f'INDEX {index}')
                                        print(f"ID COUNT {len(self.product_counter)}")
                                        print(f"handle = {self.hanle}")


                                        # remove duplicates from list
                                        print("++++++++++++")
                                        # if index % 2 == 0:
                                        try:
                                            self.save_to_xlsx()
                                        except Exception as e:
                                            print(e)
                                            print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                                        index += 1
                                        print(f"Prim INDEX = {index}")


                                    # except Exception as e:
                                    #     print(e)
                                    #     print("Error on line {}".format(sys.exc_info()[-1].tb_lineno))
                        #         if index == 5:
                        #             break
                        #     if index == 5:
                        #         break
                        # if index == 5:
                        #     break


            self.save_to_xlsx()


    def get_menu_links(self):
        url = ''
        if self.webarchive == True:
            url = self.webarchive_url+""+self.domain #+"/collections"
        print(url)
        if self.webarchive == False:
            url = self.domain

        all_categpries = []
        response = requests.get(url,timeout=60)

        soup = bs(response.text, 'html.parser')
        # print(soup)
        # quit()
        # for link in soup.find('div', class_='nav nav--combined clearfix').find_all('a'):
        # for link in soup.find('div', class_='nav nav--combined center').find_all('a'):
        #
        # for link in soup.find('div', class_='grid-item text-center large--text-right').find_all('a'):
        for link in soup.find_all('a'):
            try:
                menu_link = link.get('href')
                if menu_link.find('/collections/') != -1 and menu_link.find('/products/') == -1:
                    if menu_link not in all_categpries:
                        all_categpries.append(menu_link)
            except:
                pass

        return all_categpries


    def create_xls_file(self):
        # create shopify.xlsx file

        wb = Workbook()
        ws = wb.active
        ws.append(["id","product ID","full_link","handle","collection_handele","related_collections_handle","title","title_html","ceo_title","ceo_description","product_name","full_description","full_description_html","h2_html","description_html","bullet_points_html","bullet_points_html","bullet_points_html","published_at","created_at","vendor","type","tags","price","price_min","price_max","available","price_varies","compare_at_price","compare_at_price_max","compare_at_price_varies","requires_selling_plan","selling_plan_groups","images","featured_image","variants","option1","option2","option3","variant featured_image","variant compare_at_price","variant price"])
        wb.save("shopify.xlsx")

    def clean_duplicates(self):
        wb = load_workbook("shopify.xlsx")
        sheet = wb.worksheets[0]
        data_arr = []
        hendler_arr = []
        remove_indexs_arr = []
        hendler_origin_arr = []
        # Iterate over the rows in the sheet
        indexs = 0
        for row in sheet:
            ids = row[0].value

            if ids in data_arr:
                index_ids = data_arr.index(ids)
                print(f"Index {indexs}")
                print(f"ID {ids}")
                print(f"Duble value {data_arr[index_ids]}")
                print(f"Handle value {hendler_arr[index_ids]}")
                print(f"Dublicate index {index_ids}")
                print("Dublicate")
                print(f"ORIGIN")
                print()
                if str(row[4].value) != str(hendler_arr[index_ids]):
                    print(f"Related {hendler_arr[index_ids]  +','+ row[4].value}")
                    row[5].value = row[5].value.replace(hendler_arr[index_ids],"")

                    full_collection = hendler_arr[index_ids] + "," + row[4].value

                    row[5].value = full_collection
                    print(f"Real data {row[5].value}")

                remove_indexs_arr.append(ids)
            else:
                data_arr.append(str(ids))
                hendler_arr.append(str(row[4].value))
                hendler_origin_arr.append(str(row[4].value))
            indexs += 1


        wb.save("shopify2.xlsx")
        wb.close()

        # remove dublicates using pandas
        df = pd.read_excel('shopify2.xlsx')
        df.drop_duplicates(subset=['id'], inplace=True, keep='last')
        df.to_excel('shopify2.xlsx', index=False)
        print("Done")

        # in end of A column add product counter
        wb = load_workbook("shopify2.xlsx")
        sheet = wb.worksheets[0]

        # Iterate over the rows in the sheet
        indexs = 0
        products_arr = []
        for row in sheet:
            ids = row[1].value
            if ids not in products_arr:
                products_arr.append(ids)
            indexs += 1
        print(f"Products count {len(products_arr)}")
        sheet['A'+str(indexs+1)] = f"Products count {len(products_arr)}"
        wb.save("shopify2.xlsx")
        wb.close()

    def get_all_blog_posts(self,full_blog_link):
        print(f"full_blog_link {full_blog_link}")
        print(f"self.blog_links. {len(self.blog_links)}")
        # quit()
        try:
            respo = requests.get(full_blog_link, timeout=60)
            soup = bs(respo.text, 'html.parser')

            blog_links = soup.find_all(['a'])
            for link in blog_links:
                link = link.get('href')
                if link != None and link.find('blogs') != -1:
                    slesh_len = link.count('/')
                    if link.find('?page') != -1 and self.domain + link not in self.blog_next_pages:
                        self.blog_next_pages.append(self.domain + link)
                    if slesh_len > 2 and link.find("*/") == -1 and link.find('tagged') == -1:
                        self.blog_links.append(link)
        except:
            pass

    def requests_to_blog_posts(self,url):
        response = requests.get(url, timeout=60)
        soup = bs(response.text, 'html.parser')
        # print(soup)

        # find all links
        links = soup.find_all('a')
        blog_link = ''
        for link in links:
            # print(link)
            link = link.get('href')
            # print(link)
            if link != None and link.find('blogs') != -1:
                blog_link = link
                if blog_link.find('page') != -1 and blog_link not in self.extra_blog_pages:
                    self.extra_blog_pages.append(blog_link)
                elif blog_link not in self.blog_links:
                    print(blog_link)
                    if blog_link.find('facebook') == -1 and blog_link.find('twitter') == -1 and blog_link.find('pinterest') == -1 and blog_link.find('?page=') == -1 and link.find("*/") == -1 and link.find('tagged') == -1 and link.find('screenshot') == -1:
                        self.blog_links.append(blog_link)

    def get_blog_content(self):
        if len(self.super_webarchive_blog_links) > 0:
            self.blog_links = self.super_webarchive_blog_links

        if len(self.super_webarchive_blog_links) == 0:
            # make request to blog
            if self.webarchive == True:
                url = self.webarchive_url + self.domain + "/blogs/" + self.blog_name

            elif self.webarchive == False:
                url = self.domain + "/blogs"

            print(f"blog url {url}")
            # quit()
            self.requests_to_blog_posts(url)

                    # print(blog_link)

            # get all blogs
            full_blog_link = ''
            for extra_link in self.extra_blog_pages:
                full_blog_link = ''
                if self.webarchive == True:
                    full_blog_link = self.webarchive_url_domain + extra_link

                elif self.webarchive == False:
                    full_blog_link = self.domain + extra_link
                print(f"full_blog_link {full_blog_link}")
                self.requests_to_blog_posts(full_blog_link)
            # get all blog posts
            self.get_all_blog_posts(full_blog_link)

            for page in self.blog_next_pages:
                self.get_all_blog_posts(page)

        # get all blog links

        self.blog_links = list(dict.fromkeys(self.blog_links))
        # print(self.extra_blog_pages)
        print(self.blog_links)
        print(len(self.blog_links))
        # get all blogs data
        requests_stat = False
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            # remvoe 0 element
            # self.blog_links.pop(0)

            for link in self.blog_links:
                try:
                    if self.webarchive == True:
                        fff_link = self.webarchive_url_domain + link
                        print(fff_link)

                        page.goto(self.webarchive_url_domain + link, timeout=5000000)
                    else:
                        fff_link = self.domain + link
                        page.goto(self.domain+link, timeout=5000000)

                    # check if title loading

                    # get html content
                    html = page.content()
                    print(f"Get Html content")
                    # make soup
                    soup = bs(html, 'html.parser')

                    # get description from meta
                    try:
                        excerpt = soup.find('meta', attrs={'name': 'description'})
                        excerpt = excerpt.get('content')
                        print(f"excerpt")
                        print(excerpt)
                    except:
                        # meta property og:description
                        excerpt = soup.find('meta', property='og:description')['content']

                    try:
                        og_tags = []
                        # find tags by self.blog_tags_class
                        tags = soup.find(self.blog_div, class_=self.blog_tags_class)
                        # get all a
                        tags = tags.find_all('a')
                        for tag in tags:
                            og_tags.append(tag.text.lower())
                        og_tags = ','.join(og_tags)
                        print('===============')
                        print(og_tags)
                        print('===============')
                        # og_tags = soup.find('meta', property='og:tags')['content']
                        # print(og_tags)
                    except:
                        og_tags = ''


                    style_tags = soup.find_all(style=True)
                    for style_tag in style_tags:
                        # remove style attr
                        del style_tag['alt']
                        del style_tag['class']
                        del style_tag['style']
                        del style_tag['id']
                        del style_tag['data-mce-style']
                        del style_tag['data-mce-fragment']
                        del style_tag['data-mce-fragment']
                        del style_tag['data-mce-selected']
                        del style_tag['width']
                        del style_tag['border']
                        del style_tag['data-sheets-value']

                    all_a = soup.find_all('a')
                    for a in all_a:
                        a_href = a.get('href')
                        # remove all https://web.archive.org/ from html
                        if a_href != None:
                            if a_href.find('https://web.archive.org/') != -1:
                                a_href = a_href.replace('https://web.archive.org/','')
                                a['href'] = a_href



                    handle_pos = link.find('/blogs/')
                    if handle_pos != -1:
                        handle = link[handle_pos+7:]
                        print(handle)
                        self.blog_hendle.append(handle)

                    # get og:title
                    try:
                        og_title = soup.find('meta', property='og:title')['content']
                        print(og_title)
                        self.blog_title_text.append(og_title)
                    except:
                        self.blog_title_text.append('')

                    # get og:description
                    try:
                        ceo_desc = soup.find('meta', property='og:description')['content']
                        print(ceo_desc)
                        self.blog_desc_text.append(ceo_desc)
                    except:
                        ceo_desc = ''


                    try:
                        # get title tag
                        ceo_title = soup.find('title').text
                        print(ceo_title)
                        self.blog_ceo_title.append(ceo_title)
                    except:
                        ceo_title = ''


                    try:
                        # get title tag html
                        # find h1
                        title_html = soup.find('h1')
                        title_text = title_html.text
                        print(str(title_html))
                        print(title_text)
                        self.blog_title_html.append(str(title_html))
                    except:
                        title_text = ''
                        title_html = ''

                    try:
                        # print(soup)
                        desc_html = soup.find('div', class_='rte')
                        # print(desc_html)
                        # quit()
                        for style_tag in desc_html.find_all():
                            # remove style attr
                            del style_tag['alt']
                            del style_tag['class']
                            del style_tag['style']
                            del style_tag['id']
                            del style_tag['data-mce-style']
                            del style_tag['data-mce-fragment']
                            del style_tag['data-mce-fragment']
                            del style_tag['data-mce-selected']
                            del style_tag['width']
                            del style_tag['border']
                            del style_tag['data-sheets-value']
                        all_a = desc_html.find_all(['img','a'])
                        for a in all_a:
                            print(a.name)
                            if a.name == 'a':
                                hendl = 'href'
                            elif a.name == 'img':
                                hendl = 'src'
                            a_href = a.get(hendl)
                            print(a_href)
                            # remove all https://web.archive.org/ from html
                            if a_href != None:
                                if a_href.find('https://') != -1:
                                    # pass
                                    a_href = a_href[a_href.find('https://'):]
                                    print(a_href)
                                    a[hendl] = a_href


                        desc_text_full = desc_html.text
                        desc_html = str(desc_html).replace('""','"')
                        desc_html_full = str(desc_html)
                        print(desc_html_full)



                    except Exception as e:
                        print(e)
                        print(f"Error desc_html_full ")
                        desc_html_full = ''
                        desc_text_full = ''

                    try:
                        print(f"handle {handle}")
                        r_find = handle.rfind('/')
                        Categories = handle[:r_find]
                    except:
                        Categories = ''


                    try:
                        blog_handle = fff_link.split('/')[-1]
                    except:
                        blog_handle = ''
                    # get featured image
                    feature_image = self.feature_images(soup)
                    print(feature_image)
                    self.blog_feature_image.append(feature_image)
                    print(og_tags)
                    self.save_blog_data_to_xlsx(fff_link,blog_handle,ceo_title,Categories,ceo_desc,title_text,title_html,desc_text_full,desc_html_full,feature_image,excerpt,og_tags)
                except Exception as e:
                    print(e)




                    # get all blog content
        # self.blog_links = []
        # self.blog_hendle = []
        # self.blog_next_pages = []
        # self.blog_ceo_title = []
        # self.blog_categories = []
        # self.blog_ceo_desc = []
        # self.blog_title_text = []
        # self.blog_title_html = []
        # self.blog_desc_text = []
        # self.blog_desc_html = []
        # self.blog_feature_image = []


    def save_blog_data_to_xlsx(self,link,handle,ceo_title,Categories,ceo_desc,title_text,title_html,desc_text,desc_html,feature_image,excerpt,og_tags):
        # check if blog.xlsx exist
        try:
            wb = load_workbook("blog.xlsx")
        except:
            # if not exist create file
            wb = Workbook()
            ws = wb.active
            ws.append(["link", "handle","ceo_desc","Categories", "excerpt", "title_text", "title_html", "desc_text",
                       "desc_html","feature_image","ceo_title","og_tags"])
            wb.save("blog.xlsx")

        if title_text != 'Not Found' or ceo_title != "404 Not Found":
            print("save_blog_data_to_xlsx")
            wb = load_workbook("blog.xlsx")
            ws = wb.active
            sheet = wb.worksheets[0]

            # max row
            print(f"og_tags = {og_tags}")
            next_row = sheet.max_row + 1
            ws.cell(row=next_row, column=1, value=str(link).strip())
            ws.cell(row=next_row, column=2, value=str(handle).strip())
            ws.cell(row=next_row, column=3, value=str(excerpt).strip())
            ws.cell(row=next_row, column=4, value=str(Categories).strip())
            ws.cell(row=next_row, column=5, value=str(ceo_desc).strip())
            ws.cell(row=next_row, column=6, value=str(title_text).strip())
            ws.cell(row=next_row, column=7, value=str(title_html).strip())
            ws.cell(row=next_row, column=8, value=str(desc_text).strip())
            ws.cell(row=next_row, column=9, value=str(desc_html).strip())
            ws.cell(row=next_row, column=10, value=str(feature_image).strip())
            ws.cell(row=next_row, column=11, value=str(ceo_title).strip())
            ws.cell(row=next_row, column=12, value=str(og_tags).strip())
            next_row += 1
            wb.save("blog.xlsx")

            wb.close()

    def feature_images(self,soup):
        feature_stc = ''
        try:
            main_contant = soup.find('main')
            # find all img tags
            imgs = main_contant.find_all('img')
            for img in imgs:
                # get src
                feature_stc = img.get('src')
                print(feature_stc)
                if feature_stc.find('300x300') != -1:
                    feature_stc = feature_stc.replace('300x300','1200x1200')
                # remove /https:/
                if feature_stc.find('/https://') != -1:
                    feature_stc = feature_stc[feature_stc.find('/https:/')+1:]
                if len(feature_stc) > 5:
                    break
        except:
            feature_stc = ''
        return feature_stc

    def save_collections_data_to_xlsx(self,full_link,handle,ceo_title,ceo_description,title_text,title_html,desc_text,desc_html,col_img):
        # try:
        #     # remove collections file
        #     # os.remove("collections.xlsx")
        #     # create file if not exist
        #     wb = Workbook()
        #     ws = wb.active
        #     ws.append(["link", "handle", "ceo_title", "ceo_desc", "title_text", "title_html", "desc_text",
        #                "desc_html"])
        #     wb.save("collections.xlsx")
        # except:
        #     pass
        try:
            wb = load_workbook("collections.xlsx")
        except:
            # create file if not exist
            wb = Workbook()
            ws = wb.active
            ws.append(["link", "handle", "ceo_title", "ceo_desc", "title_text", "title_html", "desc_text",
                       "desc_html","feature_image"])
            wb.save("collections.xlsx")

        ws = wb.active
        sheet = wb.worksheets[0]

        # max row
        next_row = sheet.max_row + 1
        ws.cell(row=next_row, column=1, value=str(full_link).strip())
        ws.cell(row=next_row, column=2, value=str(handle).strip())
        ws.cell(row=next_row, column=3, value=str(ceo_title).strip())
        ws.cell(row=next_row, column=4, value=str(ceo_description).strip())
        ws.cell(row=next_row, column=5, value=str(title_text).strip())
        ws.cell(row=next_row, column=6, value=str(title_html).strip())
        ws.cell(row=next_row, column=7, value=str(desc_text).strip())
        ws.cell(row=next_row, column=8, value=str(desc_html).strip())
        ws.cell(row=next_row, column=9, value=str(col_img).strip())
        next_row += 1
        wb.save("collections.xlsx")
        wb.close()
        print(f"Save collection is done")

    def scaping_collections_data(self,all_categpries):
        if len(self.super_webarchive_collections_links) > 0:
            for category in self.super_webarchive_collections_links:
                full_link = self.webarchive_url_domain + category
                handle = category.split('/')[-1]
                print(full_link)
                col_p = category.find('/collections/')
                dm_p = category.find(self.domain)
                if col_p > dm_p and full_link.find('.json') == -1:
                    self.collection_request(full_link, handle)


        elif len(all_categpries) > 0:
            colect_index = 0
            for category in all_categpries:
                try:
                    cat_pos = category.find('/products/')
                    if cat_pos != -1:
                        category = category[:cat_pos]
                except:
                    pass
                print(category)
                print(colect_index)
                if self.webarchive == True:
                    full_link = self.webarchive_url_domain + category
                else:
                    full_link = self.domain + category
                print(full_link)
                # quit()
                handle = category.split('/')[-1]
                if full_link.find('.json') == -1:
                    self.collection_request(full_link,handle)
                    # except:
                    #     pass
                colect_index += 1

    def collection_request(self,full_link,handle):

        response = requests.get(full_link, timeout=90)
        print(response.status_code)
        if response.status_code == 200:
            print(response.status_code)
            # try:
            soup = bs(response.text, 'html.parser')
            # find all a
            links_a = soup.find_all('a')
            for link in links_a:
                if link is not None:
                    try:
                        ll = link.get('href')
                        # print(ll)
                        # find /https://
                        htt_p = ll.find('/https://')
                        resu = ll[htt_p+1:]
                        link['href'] = resu
                    except:
                        pass


            try:
                ceo_title = soup.find('title').text
                ceo_description = soup.find('meta', property='og:description')['content']
                # title_text = soup.find('title').text
            except:
                ceo_title = ''
                ceo_description = ''
                # title_text = ''
            try:
                collection_img = soup.find('meta', property='og:image')['content']
                col_pos = collection_img.find('_/http')
                col_img = collection_img[col_pos + 10:]
                print(f"col_img {col_img}")
                col_img = "http://"+col_img
            except:
                col_img = ''

            try:
                hendle_blog = full_link.split('/')[-1]
                print(f'hendle {hendle_blog}')
            except:
                hendle_blog = ''
            try:
                title_html = soup.find('h1')
                title_text = title_html.text
            except:
                title_html = ''
                title_text = ''

            try:
                desc_html = soup.find('div', class_='rte')
                desc_text = desc_html.text
            except:
                desc_html = ''
                desc_text = ''

            self.save_collections_data_to_xlsx(full_link, hendle_blog, ceo_title, ceo_description, title_text, title_html,
                                               desc_text, desc_html,col_img)

    def scrap_webarchive(self):
        url = "http://web.archive.org/web/*/"+self.domain+"*"

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()

            page.goto(url, timeout=100000)
            # page wait until id resultsUrl_info
            page.wait_for_selector('#resultsUrl_info')

            # get data by id resultsUrl_info
            soup = bs(page.content(), 'html.parser')
            res = soup.find('div', id='resultsUrl_info')
            tt = res.text
            f_p = tt.find('of')
            l_p = tt.find('entries')
            total_pages = tt[f_p+3:l_p-1]
            try:
                total_pages = int(total_pages.replace(',',''))
            except:
                total_pages = float(total_pages)
            print(total_pages)
            # print(res.text)
            all_pages = total_pages/50


            all_pages = int(all_pages)+1
            # print(all_pages)


            # get html content
            for i in range(0,all_pages):
                self.webarchive_page_par_page(page)
            print(len(self.super_webarchive_collections_links))
            print(self.super_webarchive_collections_links)
            print(len(self.super_webarchive_blog_links))
            print(self.super_webarchive_blog_links)
            print(len(self.super_webarchive_products_links))
            print(self.super_webarchive_products_links)
            # quit()

    def webarchive_page_par_page(self,page):

        html = page.content()

        # make soup
        soup = bs(html, 'html.parser')
        # find all links
        links = soup.find_all('a')
        for link in links:
            # print(link)
            link = link.get('href')
            if link != None: # and link.find(self.domain) != -1
                print(link)
                if link.find("?page=") == -1 and link.find(".oembed") == -1 and link.find(".atom") == -1:
                    try:
                        link = link.replace('*','')
                    except:
                        link = link
                    if link.find('/blogs/') != -1 and link.find('?page=') == -1:
                        # print('Done')
                        if link not in self.super_webarchive_blog_links:
                            self.super_webarchive_blog_links.append(link)
                    elif link.find('/products/') != -1:
                        if link.find('/collections/') != -1:
                            p_pos = link.find('/products/')
                            c_pos = link.find('/collections/')
                            link_space = link[c_pos:p_pos]
                            link = link.replace(link_space,'')
                        if link not in self.super_webarchive_products_links:
                            self.super_webarchive_products_links.append(link)
                    elif link.find('/collections/') != -1 and link.find('/products/') == -1: #  and link.find('facebook') == -1 and link.find('twitter') == -1 and link.find('pinterest') == -1 and link.find('?page=') == -1 and link.find("*/") == -1 and link.find('tagged') == -1 and link.find('screenshot') == -1
                        # print(f"Collection link {link}")
                        if link not in self.super_webarchive_collections_links:
                            self.super_webarchive_collections_links.append(link)

        # click on next page
        page.click('text=Next')
        print('Next page')
        time.sleep(2)
        # self.webarchive_page_par_page(page)

    def call_parent(self,res):
        print('call_parent')
        new_parent = res.parent
        return new_parent

    def find_ul_data(self,aa,related_collections):
        # get parents element of aa
        super_parent = self.call_parent(aa)
        # find ul in super_parent
        # try:
        ul = super_parent.find('ul')
        # ul is True, then remove ul and data inside
        if ul is not None:
            ul.decompose()
            # print(super_parent)
            # find a tag in super_parent
            # print(super_parent)
            a = super_parent.find('a')
            if a is not None:
                other_collection = a.get('href')
                # print(f"]]]]]]]]]]]]]]]]")
                # print(other_collection)
                # print(related_collections)
                if "/collections" in other_collection:
                    clean_collections = self.clean_collections(other_collection)
                    # print(clean_collections)
                    try:
                        clean_collections = clean_collections.replace('#','')
                    except:
                        pass

                    if len(related_collections) == 0 and len(clean_collections.strip()) > 2:
                        related_collections += clean_collections
                    elif len(clean_collections.strip()) > 4:
                        add_to_collection = ","+clean_collections
                        if len(add_to_collection) > 2:
                            related_collections += add_to_collection
                    print(f"Related collections {related_collections} and clean_collections {clean_collections} and {len(clean_collections.strip())}")
                else:
                    related_collections = self.find_ul_data(super_parent, related_collections)

        else:
            # call patern
            related_collections = self.find_ul_data(super_parent,related_collections)


        return related_collections
    def clean_collections(self,collection)->str:
        return  str(collection.split('/')[-1])


    def check_desc(self):
        # read shopify2.xlsx
        wb = load_workbook("shopify2.xlsx")
        sheet = wb.worksheets[0]
        data_arr = []
        hendler_arr = []
        for row in sheet:
            desc = row[14].value
            bullet = row[15].value
            compare_at_price = row[28].value
            compare_at_price_max = row[29].value
            related_collection = row[5].value

            img34 = row[33].value
            img35 = row[34].value
            img40 = row[39].value
            try:
                if desc.find('<ul>') !=-1 or desc.find('<ol>') != -1:
                    print(desc)
                    # print(f"{row[14].value}")
                    # print(desc)
                    # print(bullet)
                    row[14].value = ''
                    row[15].value = desc
            except:
                pass
                # wb.save("shopify2.xlsx")

            if compare_at_price == '0' or compare_at_price_max == '0':
                # print(compare_at_price)
                # print(compare_at_price_max)
                row[28].value = ''
                row[29].value = ''

            if img34:
                split_img = img34.split(',')
                for img in split_img:
                    if img.find('cdn') != -1 and img.find('http') == -1:
                        split_img[split_img.index(img)] = 'https://' + img
                # print(split_img)
                row[33].value = ",".join(split_img)

            if img35:
                split_img = img35.split(',')
                for img in split_img:
                    if img.find('cdn') != -1 and img.find('http') == -1:
                        split_img[split_img.index(img)] = 'https://' + img
                row[34].value = ",".join(split_img)
            #
            if img40:
                # print(img40)
                split_img = img40.split(',')
                for img in split_img:
                    if img.find('cdn') != -1 and img.find('http') == -1:
                        split_img[split_img.index(img)] = 'https://' + img
                # print(split_img)
                row[39].value = ",".join(split_img)

            if related_collection:
                related_collection_arr = related_collection.split(',')
                # print(related_collection_arr)
                related_collection_arr = list(dict.fromkeys(related_collection_arr))
                row[5].value = ",".join(related_collection_arr)


        wb.save("shopify2.xlsx")

    def scrap_sitemap_link(self):
        # make request to sitemap link
        url = self.domain + "/"+self.sitemap_link
        print(url)
        response = requests.get(url, timeout=90)
        print(response.status_code)
        # get data
        if response.status_code == 200:
            data = response.text
            print(type(data))

            # str to json
            data = xmltodict.parse(data)

            all_link = data['urlset']['url']

            for link in all_link:
                self.super_webarchive_products_links.append(link['loc'])

            print(len(self.super_webarchive_products_links))



if __name__ == "__main__":
    shopify_scrapper = ShopifyScrapper()
    shopify_scrapper.webarchive = True
    shopify_scrapper.webarchive_url = "http://web.archive.org/web/20230202000000/"
    shopify_scrapper.webarchive_url_domain = "http://web.archive.org"
    shopify_scrapper.blog_name = "blog-cocktails"
    shopify_scrapper.blog_tags_class = 'content-block content-block--small'
    shopify_scrapper.blog_div = 'div'

    shopify_scrapper.menu_tag = 'div'
    shopify_scrapper.menu_id = 'id'
    shopify_scrapper.menu_selector_value = 'SiteNavParent'


    shopify_scrapper.domain = "https://bonheur-tibetain.fr"
    shopify_scrapper.sitemap_link = 'sitemap_products_1.xml?from=6695409156249&to=6820545462425'
    all_categpries = []
    if shopify_scrapper.webarchive == True:
        shopify_scrapper.scrap_webarchive()

    #remove duplicates from self.super_webarchive_products_links
    print(len(list(dict.fromkeys(shopify_scrapper.super_webarchive_products_links))))
    print(shopify_scrapper.super_webarchive_products_links)



    shopify_scrapper.create_xls_file()
    if shopify_scrapper.webarchive == False:
        # all_categpries = shopify_scrapper.get_menu_links()
        # get full list of links
        shopify_scrapper.scrap_sitemap_link()


    # shopify_scrapper.scrap_shopify(all_categpries)
    # shopify_scrapper.clean_duplicates()
    # shopify_scrapper.check_desc()
    # #
    shopify_scrapper.scaping_collections_data(all_categpries)
    # get blog content data
    # shopify_scrapper.get_blog_content()



    """
    https://vintage-styles.fr - 360
    univers fleuri  - 874
    https://le-japonais-kawaii.com - 309
    https://kaneki-shop.com - 535
    """

    # .atom .oembed
    # wb = load_workbook("blog.xlsx")
    # sheet = wb.worksheets[0]
    # data_arr = []
    # hendler_arr = []
    # remove_indexs_arr = []
    # hendler_origin_arr = []
    # # Iterate over the rows in the sheet
    # indexs = 0
    # for row in sheet:
    #     # ids = row[0].value
    #     url = row[1].value
    #     hendler = row[2].value
    #     # print(ids)
    #     print(url)
    #     hh = url.split('/')[-1]
    #     print(hh)
    #     # print(hendler)
    #
    #     row[1].value = hh
    #
    #     # save to xlsx
    #     wb.save("blog.xlsx")

    # http://bonheur-tibetain.fr/
    # https://cocktail-paradis.com/










